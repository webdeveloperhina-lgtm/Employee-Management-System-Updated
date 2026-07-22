from django.shortcuts import render, redirect, get_object_or_404
from .models import Employee, Attendance, Leave
from .forms import EmployeeForm, LeaveForm
from django.contrib.auth import logout
from django.contrib.auth import authenticate
from django.contrib.auth import login
from django.contrib.auth.models import User
from django.contrib.auth.decorators import login_required
from django.db.models import Count
from django.core.paginator import Paginator
from django.http import HttpResponse
from django.views.decorators.http import require_POST
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import mm
from reportlab.lib import colors
from django.utils import timezone
from datetime import timedelta
from .models import Employee, Attendance, Leave, Task, Announcement, EmployeeDocument, PerformanceReview, AuditLog
from .forms import EmployeeForm, LeaveForm, TaskForm, AnnouncementForm, EmployeeDocumentForm, PerformanceReviewForm
from .decorators import hr_required
from .audit import record_audit

def login_view(request):

    if request.method == "POST":

        username = request.POST.get("username")
        password = request.POST.get("password")

        user = authenticate(
            request,
            username=username,
            password=password
        )

        if user:
            login(request, user)

            employee = Employee.objects.filter(user=user).first()

            if employee:
                return redirect('employee_dashboard')

            return redirect('home')

    return render(request, "login.html")


@hr_required
def home(request):

    search = request.GET.get('search')
    department = request.GET.get('department')

    employees = Employee.objects.order_by('name', 'id')

    department_data = list(
        Employee.objects
        .values('department')
        .annotate(total=Count('id'))
        .order_by('-total', 'department')
    )

    if search:
        employees = employees.filter(
            name__icontains=search
        )

    if department:
        employees = employees.filter(
            department=department
        )

    total_employees = employees.count()

    departments = (
        employees.values('department')
        .distinct()
        .count()
    )

    today = timezone.now().date()
    total_all_employees = Employee.objects.count()
    present_today = Attendance.objects.filter(date=today, status="Present").count()

    if total_all_employees > 0:
        attendance = round((present_today / total_all_employees) * 100)
    else:
        attendance = 0

    active_employees = Employee.objects.filter(status="Active").count()
    inactive_employees = total_all_employees - active_employees
    pending_leaves = Leave.objects.filter(status="Pending").count()
    approved_leaves = Leave.objects.filter(status="Approved").count()
    pending_tasks = Task.objects.filter(status="Pending").count()
    in_progress_tasks = Task.objects.filter(status="In Progress").count()
    completed_tasks = Task.objects.filter(status="Completed").count()

    attendance_labels = []
    attendance_present = []
    attendance_absent = []
    attendance_leave = []
    for offset in range(6, -1, -1):
        day = today - timedelta(days=offset)
        attendance_labels.append(day.strftime("%d %b"))
        day_records = Attendance.objects.filter(date=day)
        attendance_present.append(day_records.filter(status="Present").count())
        attendance_absent.append(day_records.filter(status="Absent").count())
        attendance_leave.append(day_records.filter(status="Leave").count())

    paginator = Paginator(employees, 5)
    page_number = request.GET.get('page')
    employees = paginator.get_page(page_number)

    context = {
        'employees': employees,
        'total_employees': total_employees,
        'departments': departments,
        'attendance': attendance,
        'department_data': department_data,
        'department_labels': [item['department'] or 'Unassigned' for item in department_data],
        'department_totals': [item['total'] for item in department_data],
        'active_employees': active_employees,
        'inactive_employees': inactive_employees,
        'present_today': present_today,
        'absent_today': Attendance.objects.filter(date=today, status="Absent").count(),
        'leave_today': Attendance.objects.filter(date=today, status="Leave").count(),
        'pending_leaves': pending_leaves,
        'approved_leaves': approved_leaves,
        'pending_tasks': pending_tasks,
        'in_progress_tasks': in_progress_tasks,
        'completed_tasks': completed_tasks,
        'total_tasks': pending_tasks + in_progress_tasks + completed_tasks,
        'attendance_labels': attendance_labels,
        'attendance_present': attendance_present,
        'attendance_absent': attendance_absent,
        'attendance_leave': attendance_leave,
        'recent_announcements': Announcement.objects.all()[:4],
        'all_departments': Employee.objects.values_list(
            'department',
            flat=True
        ).distinct(),
        'recent_activities': AuditLog.objects.select_related('actor')[:6],
    }

    return render(
        request,
        'home.html',
        context
    )


@hr_required
def employee_list(request):
    search = request.GET.get('search')
    department = request.GET.get('department')

    employees = Employee.objects.order_by('name', 'id')

    if search:
        employees = employees.filter(name__icontains=search)

    if department:
        employees = employees.filter(department=department)

    paginator = Paginator(employees, 8)
    page_number = request.GET.get('page')
    employees = paginator.get_page(page_number)

    context = {
        'employees': employees,
        'total_employees': Employee.objects.count(),
        'active_employees': Employee.objects.filter(status="Active").count(),
        'departments': Employee.objects.values_list(
            'department',
            flat=True
        ).distinct().order_by('department'),
        'selected_department': department,
    }

    return render(request, 'employee_list.html', context)


@hr_required
def department_list(request):

    department_data = (
        Employee.objects
        .values('department')
        .annotate(total=Count('id'))
        .order_by('department')
    )

    total_employees = Employee.objects.count()

    context = {
        'department_data': department_data,
        'total_employees': total_employees,
    }

    return render(request, 'departments.html', context)


@hr_required
def add_employee(request):
    if request.method == 'POST':
        form = EmployeeForm(request.POST, request.FILES)

        if form.is_valid():

            employee = form.save(commit=False)

            username = form.cleaned_data.get('username')
            password = form.cleaned_data.get('password')

            if username and password:
                user = User.objects.create_user(
                    username=username,
                    password=password,
                    email=form.cleaned_data.get('email')
                )
                employee.user = user

            employee.save()

            record_audit(
                request, "CREATE", "Employee", employee.name,
                f"Added employee {employee.name} to {employee.department}."
            )

            return redirect('home')
    else:
        form = EmployeeForm()

    return render(request, 'add_employee.html', {
        'form': form,
        'title': 'Add Employee',
        'button_text': 'Save Employee'
    })


@hr_required
def edit_employee(request, id):
    employee = get_object_or_404(Employee, id=id)

    if request.method == 'POST':
        form = EmployeeForm(
            request.POST,
            request.FILES,
            instance=employee
        )

        if form.is_valid():

            updated_employee = form.save(commit=False)

            username = form.cleaned_data.get('username')
            password = form.cleaned_data.get('password')

            if username:
                if employee.user:
                    employee.user.username = username
                    employee.user.email = form.cleaned_data.get('email')
                    if password:
                        employee.user.set_password(password)
                    employee.user.save()
                else:
                    user = User.objects.create_user(
                        username=username,
                        password=password,
                        email=form.cleaned_data.get('email')
                    )
                    updated_employee.user = user

            updated_employee.save()

            record_audit(
                request, "UPDATE", "Employee", updated_employee.name,
                f"Updated employee profile for {updated_employee.name}."
            )

            return redirect('home')
    else:
        initial_data = {}
        if employee.user:
            initial_data['username'] = employee.user.username

        form = EmployeeForm(instance=employee, initial=initial_data)

    return render(request, 'add_employee.html', {
        'form': form,
        'title': 'Edit Employee',
        'button_text': 'Update Employee'
    })


@hr_required
def delete_employee(request, id):
    employee = get_object_or_404(Employee, id=id)

    if request.method == 'POST':
        employee_name = employee.name
        employee.delete()
        record_audit(
            request, "DELETE", "Employee", employee_name,
            f"Deleted employee {employee_name}."
        )
        return redirect('home')

    return render(request, 'delete_employee.html', {
        'employee': employee
    })


def logout_view(request):

    logout(request)

    return redirect("login")


@hr_required
def employee_detail(request, id):

    employee = get_object_or_404(
        Employee,
        id=id
    )

    return render(
        request,
        'employee_detail.html',
        {
            'employee': employee
        }
    )


@hr_required
def employee_documents(request, employee_id):
    employee = get_object_or_404(Employee, id=employee_id)

    if request.method == "POST":
        form = EmployeeDocumentForm(request.POST, request.FILES)
        if form.is_valid():
            document = form.save(commit=False)
            document.employee = employee
            document.save()
            record_audit(
                request, "CREATE", "Document", document.title,
                f"Uploaded {document.document_type} for {employee.name}."
            )
            return redirect("employee_documents", employee_id=employee.id)
    else:
        form = EmployeeDocumentForm()

    return render(request, "employee_documents.html", {
        "employee": employee,
        "documents": employee.documents.all(),
        "form": form,
    })


@hr_required
def delete_employee_document(request, document_id):
    document = get_object_or_404(EmployeeDocument, id=document_id)
    employee_id = document.employee_id

    if request.method == "POST":
        document_title = document.title
        employee_name = document.employee.name
        document.file.delete(save=False)
        document.delete()
        record_audit(
            request, "DELETE", "Document", document_title,
            f"Deleted document {document_title} for {employee_name}."
        )

    return redirect("employee_documents", employee_id=employee_id)


@hr_required
def performance_reviews(request, employee_id):
    employee = get_object_or_404(Employee, id=employee_id)
    editing_review = None

    review_id = request.GET.get("edit")
    if review_id:
        editing_review = get_object_or_404(
            PerformanceReview, id=review_id, employee=employee
        )

    if request.method == "POST":
        posted_review_id = request.POST.get("review_id")
        if posted_review_id:
            editing_review = get_object_or_404(
                PerformanceReview, id=posted_review_id, employee=employee
            )

        form = PerformanceReviewForm(request.POST, instance=editing_review)
        if form.is_valid():
            review = form.save(commit=False)
            review.employee = employee
            review.reviewer = request.user
            review.save()
            action = "UPDATE" if editing_review else "CREATE"
            action_label = "Updated" if editing_review else "Created"
            record_audit(
                request, action, "Performance", review.review_period,
                f"{action_label} performance review for {employee.name}."
            )
            return redirect("performance_reviews", employee_id=employee.id)
    else:
        form = PerformanceReviewForm(instance=editing_review)

    reviews = employee.performance_reviews.select_related("reviewer")
    return render(request, "performance_reviews.html", {
        "employee": employee,
        "reviews": reviews,
        "form": form,
        "editing_review": editing_review,
    })


@hr_required
def delete_performance_review(request, review_id):
    review = get_object_or_404(PerformanceReview, id=review_id)
    employee_id = review.employee_id
    if request.method == "POST":
        review_period = review.review_period
        employee_name = review.employee.name
        review.delete()
        record_audit(
            request, "DELETE", "Performance", review_period,
            f"Deleted performance review for {employee_name} ({review_period})."
        )
    return redirect("performance_reviews", employee_id=employee_id)


@hr_required
def audit_log_list(request):
    module_filter = request.GET.get("module", "")
    action_filter = request.GET.get("action", "")
    logs = AuditLog.objects.select_related("actor")

    if module_filter:
        logs = logs.filter(module=module_filter)
    if action_filter:
        logs = logs.filter(action=action_filter)

    paginator = Paginator(logs, 15)
    page = paginator.get_page(request.GET.get("page"))

    return render(request, "audit_logs.html", {
        "logs": page,
        "modules": AuditLog.objects.values_list("module", flat=True).distinct().order_by("module"),
        "selected_module": module_filter,
        "selected_action": action_filter,
    })

@hr_required
def export_employees(request):

    search = request.GET.get('search')
    department = request.GET.get('department')

    employees = Employee.objects.all()

    if search:
        employees = employees.filter(name__icontains=search)

    if department:
        employees = employees.filter(department=department)

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Employees"

    headers = ["Name", "Email", "Department", "Status"]
    sheet.append(headers)

    header_fill = PatternFill(
        start_color="206BC4",
        end_color="206BC4",
        fill_type="solid"
    )
    header_font = Font(color="FFFFFF", bold=True)

    for col_num, header in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for employee in employees:
        sheet.append([
            employee.name,
            employee.email,
            employee.department,
            "Active"
        ])

    for column_cells in sheet.columns:
        max_length = max(
            len(str(cell.value)) if cell.value else 0
            for cell in column_cells
        )
        column_letter = column_cells[0].column_letter
        sheet.column_dimensions[column_letter].width = max_length + 6

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="employees.xlsx"'

    workbook.save(response)

    return response


@hr_required
def generate_employee_pdf(request, id):

    employee = get_object_or_404(Employee, id=id)

    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="{employee.name}_profile.pdf"'

    page = canvas.Canvas(response, pagesize=A4)
    width, height = A4

    page.setFillColor(colors.HexColor("#206bc4"))
    page.rect(0, height - 30 * mm, width, 30 * mm, fill=True, stroke=False)

    page.setFillColor(colors.white)
    page.setFont("Helvetica-Bold", 20)
    page.drawString(20 * mm, height - 18 * mm, "Employee Profile Report")

    page.setFont("Helvetica", 11)
    page.drawString(20 * mm, height - 25 * mm, "Employee Management System")

    photo_top = height - 45 * mm

    if employee.photo:
        try:
            page.drawImage(
                employee.photo.path,
                20 * mm,
                photo_top - 35 * mm,
                width=35 * mm,
                height=35 * mm,
                preserveAspectRatio=True,
                mask='auto'
            )
            details_x = 65 * mm
        except Exception:
            details_x = 20 * mm
    else:
        details_x = 20 * mm

    page.setFillColor(colors.HexColor("#111827"))
    page.setFont("Helvetica-Bold", 16)
    page.drawString(details_x, photo_top - 5 * mm, employee.name)

    page.setFont("Helvetica", 12)
    page.setFillColor(colors.HexColor("#374151"))

    details = [
        ("Email", employee.email),
        ("Department", employee.department),
        ("Status", employee.status),
    ]

    y_position = photo_top - 15 * mm

    for label, value in details:
        page.setFont("Helvetica-Bold", 11)
        page.drawString(details_x, y_position, f"{label}:")
        page.setFont("Helvetica", 11)
        page.drawString(details_x + 30 * mm, y_position, str(value))
        y_position -= 8 * mm

    page.setStrokeColor(colors.HexColor("#e5e7eb"))
    page.line(20 * mm, 20 * mm, width - 20 * mm, 20 * mm)

    page.setFont("Helvetica", 9)
    page.setFillColor(colors.HexColor("#9ca3af"))
    page.drawString(20 * mm, 14 * mm, "Generated automatically by Employee Management System")

    page.showPage()
    page.save()

    return response


@hr_required
def mark_attendance(request):

    today = timezone.now().date()
    employees = Employee.objects.all()

    todays_records = Attendance.objects.filter(date=today)

    attendance_map = {
        record.employee_id: record.status
        for record in todays_records
    }

    employee_rows = []
    for employee in employees:
        employee_rows.append({
            'employee': employee,
            'status': attendance_map.get(employee.id, None)
        })

    present_count = todays_records.filter(status="Present").count()
    absent_count = todays_records.filter(status="Absent").count()
    leave_count = todays_records.filter(status="Leave").count()

    context = {
        'employee_rows': employee_rows,
        'today': today,
        'present_count': present_count,
        'absent_count': absent_count,
        'leave_count': leave_count,
        'total_employees': employees.count(),
    }

    return render(request, 'attendance.html', context)


@hr_required
@require_POST
def mark_attendance_status(request, employee_id, status):
    valid_statuses = dict(Attendance.STATUS_CHOICES)
    if status not in valid_statuses:
        return HttpResponse("Invalid attendance status.", status=400)
    employee = get_object_or_404(Employee, id=employee_id)
    today = timezone.now().date()

    Attendance.objects.update_or_create(
        employee=employee,
        date=today,
        defaults={'status': status}
    )

    record_audit(
        request, "UPDATE", "Attendance", employee.name,
        f"Marked {employee.name} as {status} for {today}."
    )

    return redirect('mark_attendance')


@hr_required
def attendance_history(request, employee_id):

    employee = get_object_or_404(Employee, id=employee_id)
    records = Attendance.objects.filter(employee=employee)

    context = {
        'employee': employee,
        'records': records,
    }

    return render(request, 'attendance_history.html', context)


@hr_required
def apply_leave(request, employee_id):

    employee = get_object_or_404(Employee, id=employee_id)

    if request.method == 'POST':
        form = LeaveForm(request.POST)

        if form.is_valid():
            leave = form.save(commit=False)
            leave.employee = employee
            leave.save()
            return redirect('leave_list')
    else:
        form = LeaveForm()

    return render(request, 'apply_leave.html', {
        'form': form,
        'employee': employee,
        'title': 'Apply Leave',
        'button_text': 'Submit Leave Request'
    })


@hr_required
def leave_list(request):

    search = request.GET.get('search')
    status_filter = request.GET.get('status')

    leaves = Leave.objects.select_related('employee').all()

    if search:
        leaves = leaves.filter(employee__name__icontains=search)

    if status_filter:
        leaves = leaves.filter(status=status_filter)

    pending_count = Leave.objects.filter(status="Pending").count()
    approved_count = Leave.objects.filter(status="Approved").count()
    rejected_count = Leave.objects.filter(status="Rejected").count()

    paginator = Paginator(leaves, 5)
    page_number = request.GET.get('page')
    leaves = paginator.get_page(page_number)

    context = {
        'leaves': leaves,
        'pending_count': pending_count,
        'approved_count': approved_count,
        'rejected_count': rejected_count,
    }

    return render(request, 'leave_list.html', context)


@hr_required
@require_POST
def update_leave_status(request, leave_id, status):
    valid_statuses = dict(Leave.STATUS_CHOICES)
    if status not in valid_statuses or status == "Pending":
        return HttpResponse("Invalid leave status.", status=400)
    leave = get_object_or_404(Leave, id=leave_id)
    leave.status = status
    leave.save()

    record_audit(
        request, "UPDATE", "Leave", leave.employee.name,
        f"Changed {leave.employee.name}'s leave request to {status}."
    )

    return redirect('leave_list')


@hr_required
def leave_history(request, employee_id):

    employee = get_object_or_404(Employee, id=employee_id)
    records = Leave.objects.filter(employee=employee)

    context = {
        'employee': employee,
        'records': records,
    }

    return render(request, 'leave_history.html', context)


@login_required
def employee_dashboard(request):

    employee = get_object_or_404(Employee, user=request.user)

    task_records = Task.objects.filter(employee=employee)

    attendance_records = Attendance.objects.filter(employee=employee)[:10]
    leave_records = Leave.objects.filter(employee=employee)[:10]

    total_present = Attendance.objects.filter(employee=employee, status="Present").count()
    total_absent = Attendance.objects.filter(employee=employee, status="Absent").count()
    pending_leaves = Leave.objects.filter(employee=employee, status="Pending").count()

    announcements = Announcement.objects.all()[:5]

    performance_reviews = list(employee.performance_reviews.all()[:3])
    average_performance = 0
    if performance_reviews:
        average_performance = round(
            sum(review.overall_rating for review in performance_reviews) / len(performance_reviews), 1
        )

    context = {
        'employee': employee,
        'attendance_records': attendance_records,
        'leave_records': leave_records,
        'task_records': task_records,
        'total_present': total_present,
        'total_absent': total_absent,
        'pending_leaves': pending_leaves,
        'announcements': announcements,
        'performance_reviews': performance_reviews,
        'average_performance': average_performance,

    }

    return render(request, 'employee_dashboard.html', context)


@login_required
def employee_apply_leave(request):

    employee = get_object_or_404(Employee, user=request.user)

    if request.method == 'POST':
        form = LeaveForm(request.POST)

        if form.is_valid():
            leave = form.save(commit=False)
            leave.employee = employee
            leave.save()
            return redirect('employee_dashboard')
    else:
        form = LeaveForm()

    return render(request, 'apply_leave.html', {
        'form': form,
        'employee': employee,
        'title': 'Apply Leave',
        'button_text': 'Submit Leave Request'
    })


@hr_required
def create_task(request):

    if request.method == "POST":

        form = TaskForm(request.POST)

        if form.is_valid():
            task = form.save()
            record_audit(
                request, "CREATE", "Task", task.title,
                f"Assigned task {task.title} to {task.employee.name}."
            )

            return redirect('task_list')

    else:

        form = TaskForm()


    return render(
        request,
        'create_task.html',
        {
            'form': form,
            'title':'Create Task'
        }
    )

@hr_required
def task_list(request):

    search = request.GET.get('search')
    status_filter = request.GET.get('status')

    tasks = Task.objects.select_related('employee').order_by('due_date', 'id')

    if search:
        tasks = tasks.filter(employee__name__icontains=search)

    if status_filter:
        tasks = tasks.filter(status=status_filter)

    paginator = Paginator(tasks, 5)
    page_number = request.GET.get('page')
    tasks = paginator.get_page(page_number)

    return render(
        request,
        'task_list.html',
        {
            'tasks': tasks
        }
    )


@hr_required
def edit_task(request, id):

    task = get_object_or_404(Task, id=id)

    if request.method == 'POST':
        form = TaskForm(request.POST, instance=task)

        if form.is_valid():
            task = form.save()
            record_audit(
                request, "UPDATE", "Task", task.title,
                f"Updated task {task.title} for {task.employee.name}."
            )
            return redirect('task_list')
    else:
        form = TaskForm(instance=task)

    return render(request, 'create_task.html', {
        'form': form,
        'title': 'Edit Task'
    })


@hr_required
def delete_task(request, id):

    task = get_object_or_404(Task, id=id)

    if request.method == 'POST':
        task_title = task.title
        employee_name = task.employee.name
        task.delete()
        record_audit(
            request, "DELETE", "Task", task_title,
            f"Deleted task {task_title} assigned to {employee_name}."
        )
        return redirect('task_list')

    return render(request, 'delete_task.html', {
        'task': task
    })


@login_required
def employee_tasks(request):

    employee = get_object_or_404(Employee, user=request.user)

    tasks = Task.objects.filter(employee=employee)

    context = {
        'employee': employee,
        'tasks': tasks,
    }

    return render(request, 'employee_tasks.html', context)


@login_required
@require_POST
def update_task_status(request, task_id, status):

    employee = get_object_or_404(Employee, user=request.user)

    task = get_object_or_404(Task, id=task_id, employee=employee)

    valid_statuses = ["Pending", "In Progress", "Completed"]

    if status not in valid_statuses:
        return HttpResponse("Invalid task status.", status=400)

    task.status = status
    task.save(update_fields=["status"])

    return redirect('employee_dashboard')


@login_required
def update_profile_photo(request):

    employee = get_object_or_404(Employee, user=request.user)

    if request.method == 'POST' and request.FILES.get('photo'):
        employee.photo = request.FILES['photo']
        employee.save()

    return redirect('employee_dashboard')


@hr_required
def announcement_list(request):

    if request.method == 'POST':
        form = AnnouncementForm(request.POST)

        if form.is_valid():
            announcement = form.save()
            record_audit(
                request, "CREATE", "Announcement", announcement.title,
                f"Published announcement {announcement.title}."
            )
            return redirect('announcement_list')
    else:
        form = AnnouncementForm()

    announcements = Announcement.objects.all()

    return render(request, 'announcements.html', {
        'form': form,
        'announcements': announcements,
    })


@hr_required
def delete_announcement(request, id):

    announcement = get_object_or_404(Announcement, id=id)

    if request.method == 'POST':
        announcement_title = announcement.title
        announcement.delete()
        record_audit(
            request, "DELETE", "Announcement", announcement_title,
            f"Deleted announcement {announcement_title}."
        )
        return redirect('announcement_list')

    return render(request, 'delete_announcement.html', {
        'announcement': announcement
    })
